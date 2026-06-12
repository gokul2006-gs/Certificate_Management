from bson import ObjectId
from bson.errors import InvalidId
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.models import User
from django.http import JsonResponse
from django.conf import settings
from django.db import connection
from django.middleware.csrf import get_token
from django.views.decorators.csrf import ensure_csrf_cookie
from rest_framework import status
from rest_framework.decorators import api_view
from rest_framework.response import Response
import re

from django.utils import timezone

from certificates.models import Certificate
from courses.models import Course
from .models import AdminLoginLog, Student
from .permissions import admin_required, is_admin, is_student
from .serializers import StudentSerializer


def _get_client_ip(request):
    x_forwarded_for = request.META.get("HTTP_X_FORWARDED_FOR")
    if x_forwarded_for:
        return x_forwarded_for.split(",")[0].strip()
    return request.META.get("REMOTE_ADDR")


DEFAULT_PASSWORD = "Tech@123"


def _to_session_value(value):
    if isinstance(value, ObjectId):
        return str(value)
    return value


def _parse_session_object_id(value):
    if isinstance(value, ObjectId):
        return value
    if isinstance(value, str):
        try:
            return ObjectId(value)
        except (InvalidId, TypeError, ValueError):
            return None
    return None


def _next_student_id():
    last_student = (
        Student.objects.filter(student_id__startswith="TSC")
        .order_by("-student_id")
        .first()
    )
    if not last_student:
        return "TSC001"

    try:
        next_number = int(last_student.student_id.replace("TSC", "")) + 1
    except ValueError:
        next_number = Student.objects.count() + 1
    return f"TSC{next_number:03d}"


def _ensure_course():
    course, _ = Course.objects.get_or_create(
        course_name="Internship Training",
        defaults={"duration": "3 Months"},
    )
    return course


@ensure_csrf_cookie
@api_view(["GET"])
def csrf_token(request):
    return Response({"csrfToken": get_token(request)})

@api_view(["POST"])
def login_view(request):
    role = request.data.get("role", "student")

    logout(request)
    request.session.flush()
    print("REQUEST DATA:", request.data)

    if role == "admin":
        username = str(request.data.get("username", "")).strip()
        password = str(request.data.get("password", "")).strip()
        user = authenticate(
            request,
            username=username,
            password=password,
        )
        # print("AUTH RESULT:", user)
        if user and user.is_staff:
            login(request, user)
            request.session["role"] = "admin"

            log = AdminLoginLog.objects.create(
                username=user.username,
                ip_address=_get_client_ip(request),
            )
            request.session["admin_log_id"] = str(log.id)
            request.session.save()
            return Response({
                "message": "Admin login success",
                "role": "admin",
                "username": user.username,
            })
        return Response(
            {"error": "Invalid admin credentials"},
            status=status.HTTP_400_BAD_REQUEST,
        )

    student_id = str(request.data.get("student_id", "")).strip()
    password = str(request.data.get("password", "")).strip()
    if not student_id or not password:
        return Response(
            {"error": "Student ID and password are required"},
            status=status.HTTP_400_BAD_REQUEST,
        )

    try:
        student = Student.objects.get(student_id__iexact=student_id)
    except Student.DoesNotExist:
        return Response(
            {"error": "Student not found"},
            status=status.HTTP_400_BAD_REQUEST,
        )

    if student.password != password:
        return Response(
            {"error": "Invalid password"},
            status=status.HTTP_400_BAD_REQUEST,
        )

    # Explicitly create a new session
    if not request.session.session_key:
        request.session.create()
    
    request.session["role"] = "student"
    request.session["student_id"] = student.student_id
    request.session.save()
    
    return Response({
        "message": "Student login success",
        "role": "student",
        "student_id": student.student_id,
        "name": student.name,
    })


@api_view(["POST"])
def logout_view(request):
    log_id = request.session.get("admin_log_id")
    if log_id:
        object_id = _parse_session_object_id(log_id)
        if object_id is not None:
            AdminLoginLog.objects.filter(id=object_id, logout_at__isnull=True).update(
                logout_at=timezone.now()
            )
    logout(request)
    request.session.flush()
    return Response({"message": "Logged out"})


@api_view(["GET"])
def session_view(request):
    role = request.session.get("role")

    if role == "admin" and not is_admin(request):
        request.session.flush()
        return Response({
            "authenticated": False,
            "role": None,
            "student_id": None,
            "is_admin": False,
        })

    if role == "student":
        student_id = request.session.get("student_id")
        if not student_id or not Student.objects.filter(student_id__iexact=student_id).exists():
            request.session.flush()
            return Response({
                "authenticated": False,
                "role": None,
                "student_id": None,
                "is_admin": False,
            })
        return Response({
            "authenticated": True,
            "role": "student",
            "student_id": student_id,
            "is_admin": False,
        })

    if role == "admin":
        return Response({
            "authenticated": True,
            "role": "admin",
            "student_id": None,
            "is_admin": True,
        })

    return Response({
        "authenticated": False,
        "role": None,
        "student_id": None,
        "is_admin": False,
    })


@api_view(["GET"])
@admin_required
def dashboard_stats(request):
    return Response({
        "students": Student.objects.count(),
        "courses": Course.objects.count(),
        "certificates": Certificate.objects.count(),
    })


@api_view(["GET"])
@admin_required
def database_connection(request):
    try:
        db_settings = settings.DATABASES["default"]
        if db_settings.get("ENGINE") == "django_mongodb_backend":
            ping_result = connection.database.command("ping")
            result = ping_result.get("ok")
        else:
            with connection.cursor() as cursor:
                cursor.execute("SELECT 1")
                result = cursor.fetchone()

        return Response({
            "status": "connected",
            "message": "MongoDB connection is healthy.",
            "database": db_settings.get("NAME"),
            "result": result,
        })

    except Exception as e:
        return Response({
            "status": "error",
            "message": str(e),
            "database": settings.DATABASES["default"].get("NAME"),
        }, status=500)


@api_view(["POST"])
@admin_required
def upload_excel(request):
    excel_file = request.FILES.get("file")
    if not excel_file:
        return Response({"error": "No file uploaded"}, status=status.HTTP_400_BAD_REQUEST)

    try:
        from openpyxl import load_workbook
    except ImportError:
        return Response(
            {"error": "OpenPyXL is not installed. Install it to upload Excel files."},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR,
        )

    try:
        workbook = load_workbook(excel_file, read_only=True, data_only=True)
        sheet = workbook.active
    except Exception:
        return Response(
            {"error": "Invalid Excel file. Upload a valid .xlsx file."},
            status=status.HTTP_400_BAD_REQUEST,
        )

    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return Response({"error": "Excel file is empty"}, status=status.HTTP_400_BAD_REQUEST)

    headers = [str(value).strip() if value is not None else "" for value in rows[0]]
    header_map = {header.lower(): index for index, header in enumerate(headers)}

    # Normalize alternative email column names to "email"
    EMAIL_ALIASES = ["e-mail id", "e-mial id", "e-mail", "email id", "emailid", "mail", "e mail id", "email address"]
    if "email" not in header_map:
        for alias in EMAIL_ALIASES:
            if alias in header_map:
                header_map["email"] = header_map[alias]
                break
        # last resort: find any key containing "mail" or "email"
        if "email" not in header_map:
            for key in header_map.keys():
                if "mail" in key or "email" in key:
                    header_map["email"] = header_map[key]
                    break

    # Normalize alternative student ID column names to "student_id"
    STUDENT_ID_ALIASES = ["reg no", "reg. no", "reg no.", "registration no", "registration number", "regno"]
    if "student_id" not in header_map:
        for alias in STUDENT_ID_ALIASES:
            if alias in header_map:
                header_map["student_id"] = header_map[alias]
                break

    if "name" not in header_map or "email" not in header_map:
        return Response(
            {"error": f"Excel file must contain Name and Email columns. Found columns: {list(headers)}"},
            status=status.HTTP_400_BAD_REQUEST,
        )

    created = []
    updated = []
    skipped = []
    warnings = []

    # Optional student_id column (if Excel provides it).
    student_id_key = None
    for key in header_map.keys():
        k = key.replace(" ", "")
        if k in ["studentid", "student_id"]:
            student_id_key = key
            break
        if "student" in key and "id" in key:
            student_id_key = key
            break

    for row in rows[1:]:
        name_value = row[header_map["name"]] if header_map["name"] < len(row) else ""
        email_value = row[header_map["email"]] if header_map["email"] < len(row) else ""
        name = str(name_value or "").strip()
        email = str(email_value or "").strip().lower()

        if not name or not email:
            skipped.append({"email": email, "reason": "Missing name or email"})
            continue

        existing = Student.objects.filter(email=email).first()
        if existing:
            student_id_changed = False
            existing.name = name or existing.name
            if student_id_key and header_map[student_id_key] < len(row):
                desired_student_id = str(row[header_map[student_id_key]] or "").strip()
                if desired_student_id and desired_student_id != existing.student_id:
                    if not Student.objects.filter(student_id=desired_student_id).exists():
                        existing.student_id = desired_student_id
                        student_id_changed = True
                    else:
                        warnings.append({"email": email, "warning": f"student_id '{desired_student_id}' already exists; keeping existing id"})
            if student_id_changed:
                existing.save(update_fields=["name", "student_id"])
            else:
                existing.save(update_fields=["name"])
            updated.append(existing.student_id)
            continue

        desired_student_id = None
        if student_id_key and header_map[student_id_key] < len(row):
            desired_student_id = str(row[header_map[student_id_key]] or "").strip()
            if not desired_student_id:
                desired_student_id = None

        if desired_student_id and Student.objects.filter(student_id=desired_student_id).exists():
            warnings.append({"email": email, "warning": f"student_id '{desired_student_id}' already exists; generating new id"})
            desired_student_id = None

        final_student_id = desired_student_id or _next_student_id()
        student = Student.objects.create(
            student_id=final_student_id,
            name=name,
            email=email,
            password=DEFAULT_PASSWORD,
        )
        created.append(student.student_id)

    return Response({
        "message": "Students uploaded successfully",
        "created_count": len(created),
        "created_student_ids": created,
        "updated_count": len(updated),
        "updated_student_ids": updated,
        "skipped": skipped,
        "warnings": warnings,
        "default_password": DEFAULT_PASSWORD,
    })


@api_view(["GET", "POST"])
@admin_required
def students(request):
    if request.method == "GET":
        student_rows = Student.objects.order_by("student_id")
        return Response(StudentSerializer(student_rows, many=True).data)

    data = request.data.copy()
    data.setdefault("student_id", _next_student_id())
    data.setdefault("password", DEFAULT_PASSWORD)

    serializer = StudentSerializer(data=data)
    if serializer.is_valid():
        serializer.save()
        return Response(serializer.data, status=status.HTTP_201_CREATED)
    return Response(
        {"error": "Could not create student", "detail": serializer.errors},
        status=status.HTTP_400_BAD_REQUEST,
    )


@api_view(["GET", "PUT", "DELETE"])
def student_detail(request, student_id):
    # Normalize student_id
    student_id = student_id.strip()
    
    try:
        student = Student.objects.get(student_id__iexact=student_id)
    except Student.DoesNotExist:
        return Response({"error": "Student not found"}, status=status.HTTP_404_NOT_FOUND)

    if request.method == "GET":
        session_student_id = request.session.get("student_id")
        if not is_admin(request) and session_student_id != student.student_id:
            return Response({"error": "Access denied"}, status=status.HTTP_403_FORBIDDEN)
        return Response(StudentSerializer(student).data)

    if not is_admin(request):
        return Response({"error": "Admin access required"}, status=status.HTTP_403_FORBIDDEN)

    if request.method == "DELETE":
        student.delete()
        return Response({"message": "Student deleted"})

    if request.method == "PUT":
        serializer = StudentSerializer(student, data=request.data, partial=True)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


@api_view(["POST"])
@admin_required
def bulk_delete_students(request):
    student_ids = request.data.get("student_ids", [])
    if not isinstance(student_ids, list) or not student_ids:
        return Response(
            {"error": "student_ids must be a non-empty list"},
            status=status.HTTP_400_BAD_REQUEST,
        )

    students_qs = Student.objects.filter(student_id__in=student_ids)
    deleted_count = students_qs.count()
    students_qs.delete()
    return Response(
        {
            "message": "Students deleted",
            "deleted_count": deleted_count,
        }
    )


@api_view(["GET"])
def student_profile(request, student_id):
    # Normalize student_id
    student_id = student_id.strip()
    
    try:
        student = Student.objects.get(student_id__iexact=student_id)
    except Student.DoesNotExist:
        return Response({"error": "Student not found"}, status=status.HTTP_404_NOT_FOUND)

    if not is_admin(request):
        session_student_id = request.session.get("student_id")
        if not is_student(request) or session_student_id != student.student_id:
            return Response({"error": "Access denied"}, status=status.HTTP_403_FORBIDDEN)

    return Response(StudentSerializer(student).data)


@api_view(["GET"])
@admin_required
def admin_login_logs(request):
    logs = AdminLoginLog.objects.all()[:100]
    data = [
        {
            "id": str(log.id),
            "username": log.username,
            "login_at": log.login_at,
            "logout_at": log.logout_at,
            "ip_address": log.ip_address,
        }
        for log in logs
    ]
    return Response(data)
@api_view(["GET"])
@admin_required
def test_env(request):
    return Response({"test": "working"})

@api_view(["GET"])
@admin_required
def admin_check(request):
    return Response(
        list(
            User.objects.values(
                "username",
                "is_staff",
                "is_superuser"
            )
        )
    )